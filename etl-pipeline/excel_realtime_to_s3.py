
''' 
This script is witten for updating a file in local directory to AWS S3 in real-time.
An existing excel file in a local computer has frequent updates. It is required that the updates are 
reflected in AWS S3 and DynamoDB in real-time. 
The updates are made in the excel file, each time the excel file is saved with new updates, the excel file 
is edited with the addition of a new ID column and converted to csv format. The updates are uploaded to S3 
in real-time, and then sent to DynamoDB table by a lambda function triggerd by CloudWatch event 
(the lambda function is shown in a separate script).
'''

import openpyxl as xl
from pathlib import Path
import csv
import boto3
import os
import shutil


# This function reads excel xlsx worksheets and then edits the files as needed.
# In this application it is required to add a new column the existing excel worksheet which serve as
# test ID's.
def edit_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb["Sheet1"]

    sheet.insert_cols(1)

    sheet.cell(1, 1).value = "Test_ID"
    for row in range(2, sheet.max_row+1):
        cell = sheet.cell(row, 2)
        test_id = str(filename)[-9: -5] + "_" + str(cell.value)
        test_id_cell = sheet.cell(row, 1)
        test_id_cell.value = test_id

    wb.save(filename)


# This function is to convert xlsx file to csv format
def to_csv(filename):
    wb = xl.load_workbook(filename)
    # sh = wb.get_active_sheet()
    sh = wb["Sheet1"]
    with open(filename, 'w', newline="") as f:  # in python 2: open('test.csv', 'wb') as f
        c = csv.writer(f, delimiter=',')
        for r in sh.rows:
            c.writerow([cell.value for cell in r])


# First copy the excel file to another directory, then edit it and convert it to csv in that directory
file_excel = Path('/Users/binggangliu/downloads/mtest_excel/MembraneTest.xlsx')
path_csv = Path('/Users/binggangliu/downloads/mtest_csv/')
shutil.copy(file_excel, path_csv)

for file in path_csv.glob('*'):
    print(file)
    if str(file) == '/Users/binggangliu/downloads/mtest_csv/MembraneTest.xlsx':
        # skip the system generated files such as ., .. and .DS_Store
        edit_workbook(file)
        to_csv(file)
        os.rename(file, str(file).replace('.xlsx', '.csv'))
    else:
        pass


# upload file with updates to S3

s3 = boto3.client('s3')
bucket_name = 'bl-csv-07012019'

for file in path_csv.glob('*'):
    if str(file) == '/Users/binggangliu/downloads/mtest_csv/MembraneTest.csv':
         # skip the system generated files such as ., .. and .DS_Store
        filenameins3 = str(file)[-16:]
#        print(filenameins3)
        content = open(file, 'rb')
        s3.put_object(
            Bucket=bucket_name,
            Key=filenameins3,
            Body=content
        )

print("Completed updated in S3!")
