
''' 
Using the Python library openpyxl to read excel xlsx files, edit the files
and then write to csv format (Pandas library is another option for this purpose). 

The transformed csv files were first saved in the local file directory as a copy, 
and then loaded to AWS S3 and DynamoDB with boto3. 
'''

import openpyxl as xl
from pathlib import Path
import csv
import boto3
from os import rename


# The function below was created to read excel xlsx worksheets and then edits them as required.
# In this application there are a series of excel files in which each holds the test results
# of the same products in different years. The requirement is to add a new column with the
# testing year added to the product ID's.
def edit_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb["Sheet1"]

    sheet.insert_cols(1)

    sheet.cell(1, 1).value = "Test_ID"
    for row in range(2, sheet.max_row+1):
        cell = sheet.cell(row, 2)
        test_id = str(filename)[-9: -5] + "_" + str(cell.value)
        test_id_cell = sheet.cell(row, 1)
        test_id_cell.value = test_id

    wb.save(filename)


# This function is to convert xlsx files into csv format:
def to_csv(filename):
    wb = xl.load_workbook(filename)
    # sh = wb.get_active_sheet()
    sh = wb["Sheet1"]
    with open(filename, 'w', newline="") as f:  # in python 2: open('test.csv', 'wb') as f
        c = csv.writer(f, delimiter=',')
        for r in sh.rows:
            c.writerow([cell.value for cell in r])


# The newly transformed csv files are saved to the local directory with 'csv' file extension:
path = Path('~/MembraneTests/')

for file in path.glob('*'):
    edit_workbook(file)
    to_csv(file)
    rename(file, str(file).replace('.xlsx', '.csv'))


# Upload csv files to S3
s3 = boto3.client('s3')
bucket_name = 'bl-csv-01-2018'

for file in path.glob('*'):
    filenameins3 = str(file)[-20: -4] + '.csv'
    print(filenameins3)
    content = open(file, 'rb')
    s3.put_object(
        Bucket=bucket_name,
        Key=filenameins3,
        Body=content
    )

print("Files successfully loaded to S3")
